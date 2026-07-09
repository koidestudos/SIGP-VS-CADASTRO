#!/usr/bin/env python3
"""Gera programacoes-viagens-xlsx.json a partir da planilha PROGRAMAÇÃO DE VIAGENS.xlsx"""
import json
import re
import sys
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl

MONTHS = {
    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'MARCO': 3, 'ABRIL': 4,
    'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9,
    'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12,
}
GERENCIA_SHEETS = ('GAS', 'GAP', 'GVS')
PROGRAMADA_COLORS = {
    'FFFFFF', 'F3F3F3', 'F2F2F2', 'F1EFC1', 'D9E1F2', 'FBD4B4',
    'FFE599', 'FFD966', 'F9CB9C', 'E06666', 'FFFF00', 'FFF000',
}


def get_fill_rgb(cell):
    fill = cell.fill
    if not fill or fill.patternType in (None, 'none'):
        return 'FFFFFF'
    for color in (fill.fgColor, fill.start_color):
        if color is None:
            continue
        if color.type == 'rgb' and color.rgb:
            rgb = color.rgb
            if len(rgb) == 8:
                rgb = rgb[2:]
            return rgb.upper()
    return 'FFFFFF'


def status_from_color(rgb):
    if rgb == '76A5AF':
        return None
    if rgb in PROGRAMADA_COLORS or rgb.startswith('theme') or rgb.startswith('idx'):
        return 'Programada'
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        if g >= 100 and g > r + 20 and g > b + 10:
            return 'Aprovado'
    except ValueError:
        pass
    return 'Programada'


def is_month_row(val):
    if not val or not isinstance(val, str):
        return False
    v = val.strip().upper().replace('Ç', 'C')
    return v in MONTHS


def month_num(val):
    v = val.strip().upper().replace('Ç', 'C')
    return MONTHS.get(v, 6)


def is_week_header(val):
    if not val:
        return False
    s = str(val).strip().upper()
    return 'SEMANA' in s and len(s) < 30


def is_semana_cell(val):
    if not val:
        return False
    s = str(val).strip()
    return bool(re.match(r'^\d+[ªº°oO]?$', s)) or bool(re.match(r'^\d+[ªº°]?\s*SEMANA', s, re.I))


def parse_bool(v):
    s = str(v or '').lower()
    return 'sim' in s and 'não' not in s and 'nao' not in s


def infer_tipo(t):
    t = (t or '').lower()
    rules = [
        ('supervis', 'Supervisão'), ('oficina', 'Capacitação'), ('capacit', 'Capacitação'),
        ('qualific', 'Capacitação'), ('reuni', 'Reunião de planejamento'),
        ('visita', 'Visita técnica'), ('vistoria', 'Visita técnica'),
        ('campanha', 'Palestra / Oficina'), ('monitor', 'Monitoramento'),
        ('semin', 'Palestra / Oficina'), ('fórum', 'Palestra / Oficina'),
        ('forum', 'Palestra / Oficina'), ('caravana', 'Ação de campo'),
    ]
    for kw, val in rules:
        if kw in t:
            return val
    return 'Visita técnica'


def fmt_date(y, m, d):
    return f'{y:04d}-{m:02d}-{d:02d}'


def parse_dates(val, month_n, default_year=2026):
    if val is None or val == '':
        return fmt_date(default_year, month_n, 1), fmt_date(default_year, month_n, 1)
    if isinstance(val, datetime):
        d = val.date().isoformat()
        return d, d
    if isinstance(val, date):
        return val.isoformat(), val.isoformat()
    s = str(val).strip()
    ym = re.search(r'(20\d{2})', s)
    year = int(ym.group(1)) if ym else default_year
    m = re.search(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](20\d{2})', s)
    if m:
        d1 = fmt_date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        return d1, d1
    m = re.search(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2})', s)
    if m:
        yr = 2000 + int(m.group(3))
        d1 = fmt_date(yr, int(m.group(2)), int(m.group(1)))
        return d1, d1
    m = re.search(r'(\d{1,2})\s*[aAàÀeE]\s*(\d{1,2})[/\-.](\d{2,4})', s)
    if m:
        d1, d2, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yr < 100:
            yr += 2000
        return fmt_date(yr, month_n, d1), fmt_date(yr, month_n, d2)
    m = re.search(r'(\d{1,2})\s*[aAàÀeE]\s*(\d{1,2})', s)
    if m:
        d1, d2 = int(m.group(1)), int(m.group(2))
        if d2 < d1 and month_n == 6:
            return fmt_date(year, 6, d1), fmt_date(year, 7, d2)
        return fmt_date(year, month_n, d1), fmt_date(year, month_n, d2)
    m = re.search(r'(\d{1,2})[/\-.](\d{1,2})', s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if a > 12:
            day, mon = a, b
        elif b > 12:
            day, mon = b, a
        else:
            day, mon = a, month_n
        return fmt_date(year, mon, day), fmt_date(year, mon, day)
    m = re.search(r'(\d{1,2})\s+de\s+\w+\s+de\s+(20\d{2})', s, re.I)
    if m:
        return fmt_date(int(m.group(2)), month_n, int(m.group(1))), fmt_date(int(m.group(2)), month_n, int(m.group(1)))
    return fmt_date(year, month_n, 1), fmt_date(year, month_n, 1)


COORD_RULES = [
    (r'casai|idoso', 'gas-idoso'),
    (r'casm|mulher|samvvis|savvis', 'gas-mulher'),
    (r'criança|adolescente|caca|casca', 'gas-crianca'),
    (r'ist|aids|cta|hiv|sífilis|htlv|hansen|tubercul|sinan|transmiss', 'gas-dt'),
    (r'capd|defici', 'gas-pcd'),
    (r'equidade|cetespi|lgbt|indígen|indigen|warao', 'gas-equidade'),
    (r'epidemiolog|cepi', 'gvs-epi'),
    (r'ambiental|cvsa', 'gvs-cvsa'),
    (r'imuniz', 'gvs-imuno'),
    (r'análise|analise|sinasc', 'gvs-analise'),
    (r'pvt', 'gvs-pvt'),
    (r'bucal|odonto', 'gap-bucal'),
    (r'planifica|aps|gaps|atenção prim|atencao prim|primaria', 'gap-aps'),
    (r'\bgas\b', 'gas-pcd'),
]


def map_coord(area, gerencia):
    text = f'{area or ""} {gerencia}'.lower()
    for pat, cid in COORD_RULES:
        if re.search(pat, text, re.I):
            return cid
    return {'GAS': 'gas-crianca', 'GAP': 'gap-aps', 'GVS': 'gvs-epi'}[gerencia]


def find_municipio(local, municipios):
    loc = (local or '').lower()
    if not loc or any(x in loc for x in ('virtual', 'online', 'remot')):
        return 'm-teresina'
    best_id = ''
    best_len = 0
    for m in municipios:
        nome = m['nome'].lower()
        if nome in loc and len(nome) > best_len:
            best_id = m['id']
            best_len = len(nome)
    if best_id:
        return best_id
    if 'teresina' in loc:
        return 'm-teresina'
    return ''


def parse_workbook(xlsx_path, geo_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_colors = openpyxl.load_workbook(xlsx_path, data_only=False)
    municipios = json.loads(Path(geo_path).read_text(encoding='utf-8'))['municipios']
    rows = []
    idx = 0
    for sheet_name in GERENCIA_SHEETS:
        ws = wb[sheet_name]
        wsc = wb_colors[sheet_name]
        current_month = ''
        current_month_num = 6
        current_week = ''
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if is_month_row(a):
                current_month = str(a).strip().upper()
                current_month_num = month_num(current_month)
                continue
            if is_week_header(a):
                current_week = str(a).strip()
                continue
            if not b or not str(b).strip():
                continue
            if str(b).strip().upper().startswith('TIPO DE'):
                continue
            status = status_from_color(get_fill_rgb(wsc.cell(r, 2)))
            if status is None:
                continue
            semana = str(a).strip() if is_semana_cell(a) else current_week
            if semana and 'semana' not in semana.lower() and re.match(r'^\d', semana):
                semana = f"{semana.replace('°', 'ª')} Semana"
            area = ws.cell(r, 3).value
            titulo = re.sub(r'\s+', ' ', str(b)).strip()
            local = str(ws.cell(r, 7).value or '').strip()
            d1, d2 = parse_dates(ws.cell(r, 6).value, current_month_num)
            mid = find_municipio(local, municipios)
            regional = next((m['regionalId'] for m in municipios if m['id'] == mid), '')
            equipe_raw = str(ws.cell(r, 10).value or '')
            idx += 1
            rows.append({
                'id': f'xls-{sheet_name.lower()}-{idx:04d}',
                'titulo': titulo[:200],
                'tipoAtividade': infer_tipo(titulo),
                'coordenacaoId': map_coord(str(area or ''), sheet_name),
                'responsavel': equipe_raw.split(',')[0].split('\n')[0][:80] or str(area or 'Equipe técnica')[:80],
                'objetivo': titulo,
                'publicoAlvo': str(ws.cell(r, 4).value or '').strip(),
                'semana': semana or current_week or current_month,
                'dataInicial': d1,
                'dataFinal': d2,
                'duracao': str(ws.cell(r, 5).value or '').strip() or '—',
                'regionalId': regional,
                'municipioId': mid,
                'localAtividade': local or 'A definir',
                'necessitaTransporte': parse_bool(ws.cell(r, 9).value),
                'necessitaAlimentacao': parse_bool(ws.cell(r, 8).value),
                'obsLogistica': '',
                'equipe': [
                    {'nome': n.strip()[:60], 'cargo': 'Integrante'}
                    for n in re.split(r'[,/\n]', equipe_raw) if len(n.strip()) > 2
                ][:10],
                'codigoOrcamentario': str(ws.cell(r, 11).value or '').strip(),
                'fonteRecurso': str(ws.cell(r, 12).value or '').strip(),
                'observacoes': f'Gerência {sheet_name} | {current_month} | Área: {area}',
                'status': status,
            })
    return rows


def main():
    root = Path(__file__).resolve().parents[1]
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else root / 'data' / 'PROGRAMACAO-DE-VIAGENS.xlsx'
    out = root / 'src' / 'data' / 'programacoes-viagens-xlsx.json'
    geo = root / 'src' / 'data' / 'regions-municipios.json'
    rows = parse_workbook(xlsx, geo)
    out.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
    stats = Counter(r['status'] for r in rows)
    print(f'Gerado {out} — {len(rows)} programações ({dict(stats)})')


if __name__ == '__main__':
    main()
